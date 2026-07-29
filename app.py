import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="YVF Adoption Dashboard – CS HAD",layout="wide")
wb=load_workbook("data/YVF_Data.xlsx",data_only=True)
ws=wb["YVF Status"]

eligible=ws["A4"].value or 0
onboard=ws["C4"].value or 0
pending=ws["E4"].value or 0
active=ws["F4"].value or 0
bookings=ws["G4"].value or 0
avg=ws["H4"].value or 0
target=ws["J4"].value or 0
adoption=(active/onboard*100) if onboard else 0
onboard_rate=(onboard/eligible*100) if eligible else 0
ach=(bookings/target*100) if target else 0

st.title("YVF Adoption Dashboard – CS HAD")
c=st.columns(4)
c[0].metric("Eligible",eligible)
c[1].metric("Onboarded",onboard)
c[2].metric("Active",active)
c[3].metric("Pending",pending)
c=st.columns(4)
c[0].metric("Adoption Rate",f"{adoption:.1f}%")
c[1].metric("Onboarding Rate",f"{onboard_rate:.1f}%")
c[2].metric("Bookings",bookings)
c[3].metric("Achievement",f"{ach:.1f}%")
st.info("Scaffold V6. Add charts/pages next.")
